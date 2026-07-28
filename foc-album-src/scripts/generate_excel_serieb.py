import os
import re
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 1. Carrega as variáveis do .env
env = {}
with open(".env", "r") as f:
    for line in f:
        if "=" in line:
            parts = line.strip().split("=")
            if len(parts) >= 2:
                env[parts[0].strip()] = "=".join(parts[1:]).strip()

url = env["VITE_SUPABASE_URL"]
headers = {
    "apikey": env["VITE_SUPABASE_ANON_KEY"],
    "Authorization": f"Bearer {env['VITE_SUPABASE_ANON_KEY']}",
}

print("Iniciando coleta de dados do Supabase...")

HOUSES = [
  'BRB', 'DIS', 'LGS', 'MRS', 'SCT', 'RDP', 'SHW', 'UNT', 'SAU', 'STA', 'UNF', 'EKW', 'GST', 'SKB'
]

HOUSE_NAMES = {
  'BRB': 'Brobnar',
  'DIS': 'Dis',
  'LGS': 'Logos',
  'MRS': 'Mars',
  'SCT': 'Sanctum',
  'RDP': 'Redemption',
  'SHW': 'Shadows',
  'UNT': 'Untamed',
  'SAU': 'Saurian',
  'STA': 'Star Alliance',
  'UNF': 'Unfathomable',
  'EKW': 'Ekwidon',
  'GST': 'Geistoid',
  'SKB': 'Skyborn'
}

# 2. Busca dados das tabelas
players = requests.get(f"{url}/rest/v1/foc2026_players?select=*", headers=headers).json()
matches = requests.get(f"{url}/rest/v1/foc2026_matches?select=*&order=round_number.asc", headers=headers).json()
collections = requests.get(f"{url}/rest/v1/foc2026_collections?select=*", headers=headers).json()
challenges = requests.get(f"{url}/rest/v1/foc2026_challenges?select=*", headers=headers).json()
stickers_log = requests.get(f"{url}/rest/v1/foc2026_stickers_log?select=*&order=created_at.desc", headers=headers).json()

# Mapeia players e filtra para Série B
b_players = [p for p in players if p.get("serie") == "B" and p["username"] not in ['teste_1', 'teste_2', 'admin', 'album']]
b_usernames = {p["username"] for p in b_players}

print(f"Encontrados {len(b_players)} jogadores na Série B.")

# Mapeia as figurinhas (de stickers.js ou dinamicamente das coleções)
# Para evitar carregar o arquivo JS, podemos montar a lista de figurinhas das coleções
all_sticker_ids = sorted(list({c["sticker_id"] for c in collections}))

# 3. Reconstrói coleções até a Rodada 6
# Identifica figurinhas obtidas na rodada 7 ou superior
stickers_to_subtract = {} # (player, sticker_id) -> quantity to subtract
sticker_pattern = re.compile(r'\b([A-Z]{3} \d+)\b')

for log in stickers_log:
    r_num = log.get("round_number")
    if r_num and r_num >= 7:
        user = log["player_username"]
        msg = log["message"]
        matches_found = sticker_pattern.findall(msg)
        for sid in matches_found:
            key = (user, sid)
            stickers_to_subtract[key] = stickers_to_subtract.get(key, 0) + 1

# Monta a coleção de cada jogador na Rodada 6
collections_r6 = {}
for p in b_players:
    collections_r6[p["username"]] = {}

for c in collections:
    user = c["player_username"]
    if user in b_usernames:
        sid = c["sticker_id"]
        qty = c["quantity"]
        # Subtrai o que foi obtido na rodada 7+
        sub = stickers_to_subtract.get((user, sid), 0)
        final_qty = max(0, qty - sub)
        if final_qty > 0:
            collections_r6[user][sid] = final_qty

# 4. Calcula classificação até a rodada 6
standings = []
for p in b_players:
    username = p["username"]
    # Quantidade de figurinhas de jogador (não terminam em ' 0')
    stickers_count = sum(1 for sid, qty in collections_r6[username].items() if not sid.endswith(" 0") and qty > 0)
    
    played = 0
    wins = 0
    losses = 0
    keys = 0
    
    for m in matches:
        if m["completed"] and m["round_number"] <= 6:
            if m["player_a_username"] == username:
                played += 1
                keys += m["player_a_keys"]
                if m["player_a_keys"] > m["player_b_keys"]:
                    wins += 1
                elif m["player_a_keys"] < m["player_b_keys"]:
                    losses += 1
            elif m["player_b_username"] == username:
                played += 1
                keys += m["player_b_keys"]
                if m["player_b_keys"] > m["player_a_keys"]:
                    wins += 1
                elif m["player_b_keys"] < m["player_a_keys"]:
                    losses += 1

    # Desafios completados até a rodada 6
    # Como não temos round_number nos desafios, usamos todos concluídos (a maioria foi feita até a rodada 6)
    completed_challenges = sum(1 for c in challenges if c["player_username"] == username and c["completed"])
    
    standings.append({
        "username": username,
        "name": p["name"],
        "stickers_count": stickers_count,
        "wins": wins,
        "losses": losses,
        "played": played,
        "keys": keys,
        "challenges": completed_challenges
    })

# Ordenação oficial: Figurinhas desc -> Vitórias desc -> Chaves desc -> Desafios desc -> Nome asc
standings.sort(key=lambda x: (-x["stickers_count"], -x["wins"], -x["keys"], -x["challenges"], x["name"]))

# 5. Prepara os dados para o openpyxl
wb = Workbook()

# Estilos Visuais Premium
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Azul Escuro
body_font = Font(name="Calibri", size=11, color="000000")
title_font = Font(name="Calibri", size=14, bold=True, color="1F4E78")
zebra_fill = PatternFill(start_color="F2F6F9", end_color="F2F6F9", fill_type="solid") # Azul claro suave

thin_side = Side(border_style="thin", color="D9D9D9")
thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

align_center = Alignment(horizontal="center", vertical="center")
align_left = Alignment(horizontal="left", vertical="center")
align_right = Alignment(horizontal="right", vertical="center")

# --- ABA 1: CLASSIFICAÇÃO ---
ws_st = wb.active
ws_st.title = "Classificação Série B (R6)"
ws_st.views.sheetView[0].showGridLines = True

ws_st.append([])
ws_st.cell(row=2, column=2, value="Classificação Oficial - Série B (Até a Rodada 6)").font = title_font
ws_st.append([])

headers_st = ["Pos", "Jogador", "Figurinhas do Álbum", "Vitórias", "Derrotas", "Jogos", "Chaves Forjadas", "Desafios Completados"]
ws_st.append([]) # Linha vazia
row_num = ws_st.max_row
for col_idx, text in enumerate(headers_st, start=2):
    cell = ws_st.cell(row=row_num, column=col_idx, value=text)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = align_center
    cell.border = thin_border

for idx, p in enumerate(standings, start=1):
    row_data = [idx, p["name"], p["stickers_count"], p["wins"], p["losses"], p["played"], p["keys"], p["challenges"]]
    ws_st.append([]) # Adiciona linha para preencher pelas colunas corretas
    r_idx = ws_st.max_row
    is_zebra = idx % 2 == 0
    for col_idx, val in enumerate(row_data, start=2):
        cell = ws_st.cell(row=r_idx, column=col_idx, value=val)
        cell.font = body_font
        cell.border = thin_border
        if is_zebra:
            cell.fill = zebra_fill
        
        # Alinhamento
        if col_idx in [2, 4, 5, 6, 7, 8, 9]:
            cell.alignment = align_center
        else:
            cell.alignment = align_left

# Ajusta largura das colunas
for col in ws_st.columns:
    max_len = max(len(str(cell.value or '')) for cell in col)
    col_letter = get_column_letter(col[0].column)
    ws_st.column_dimensions[col_letter].width = max(max_len + 3, 10)


# --- ABA 2: PARTIDAS ---
ws_ma = wb.create_sheet(title="Confrontos Série B (R1-R6)")
ws_ma.views.sheetView[0].showGridLines = True

ws_ma.append([])
ws_ma.cell(row=2, column=2, value="Histórico de Confrontos - Série B (Rodadas 1 a 6)").font = title_font
ws_ma.append([])

headers_ma = ["Rodada", "Jogador A", "Chaves A", "vs", "Chaves B", "Jogador B", "Status", "Picks Jogador A", "Picks Jogador B"]
ws_ma.append([])
row_num = ws_ma.max_row
for col_idx, text in enumerate(headers_ma, start=2):
    cell = ws_ma.cell(row=row_num, column=col_idx, value=text)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = align_center
    cell.border = thin_border

# Filtra partidas da Série B até a rodada 6
b_matches = []
for m in matches:
    if m["round_number"] <= 6:
        # Verifica se pelo menos um dos jogadores é da Série B para incluir (e exclui teste/admin)
        if m["player_a_username"] in b_usernames or m["player_b_username"] in b_usernames:
            b_matches.append(m)

for idx, m in enumerate(b_matches, start=1):
    # Busca nomes legíveis
    p_a_name = next((p["name"] for p in players if p["username"] == m["player_a_username"]), m["player_a_username"])
    p_b_name = next((p["name"] for p in players if p["username"] == m["player_b_username"]), m["player_b_username"])
    
    status = "Concluída" if m["completed"] else "Pendente"
    row_data = [
        f"Rodada {m['round_number']}",
        p_a_name,
        m["player_a_keys"] if m["player_a_reported"] else 0,
        "x",
        m["player_b_keys"] if m["player_b_reported"] else 0,
        p_b_name,
        status,
        m["player_a_picks"] or "",
        m["player_b_picks"] or ""
    ]
    
    ws_ma.append([])
    r_idx = ws_ma.max_row
    is_zebra = idx % 2 == 0
    for col_idx, val in enumerate(row_data, start=2):
        cell = ws_ma.cell(row=r_idx, column=col_idx, value=val)
        cell.font = body_font
        cell.border = thin_border
        if is_zebra:
            cell.fill = zebra_fill
        
        # Alinhamento
        if col_idx in [2, 4, 5, 6, 8]:
            cell.alignment = align_center
        else:
            cell.alignment = align_left

# Ajusta largura das colunas da aba partidas
for col in ws_ma.columns:
    max_len = max(len(str(cell.value or '')) for cell in col)
    col_letter = get_column_letter(col[0].column)
    ws_ma.column_dimensions[col_letter].width = max(min(max_len + 3, 50), 10) # limita máximo para evitar colunas de picks gigantescas


# --- ABA 3: BRASÕES E CASAS ---
ws_co = wb.create_sheet(title="Brasões Série B")
ws_co.views.sheetView[0].showGridLines = True

ws_co.append([])
ws_co.cell(row=2, column=2, value="Resumo de Brasões Obtidos - Série B (Até a Rodada 6)").font = title_font
ws_co.append([])

headers_co = ["Jogador", "Brasão", "Status", "Origem/Rodada"]
ws_co.append([])
row_num = ws_co.max_row
for col_idx, text in enumerate(headers_co, start=2):
    cell = ws_co.cell(row=row_num, column=col_idx, value=text)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = align_center
    cell.border = thin_border

# Lista os brasões de cada jogador até a rodada 6
row_idx = 1
for idx_p, p in enumerate(sorted(b_players, key=lambda x: x["name"]), start=1):
    username = p["username"]
    owned_stickers = collections_r6[username]
    
    # Brasões
    owned_emblems = sorted([sid.split(' ')[0] for sid in owned_stickers.keys() if sid.endswith(' 0')])
    
    for h in HOUSES:
        has_emblem = h in owned_emblems
        row_data = [
            p["name"],
            h,
            "Possui" if has_emblem else "Não Possui",
            "Coleção / Conquistado" if has_emblem else "-"
        ]
        
        ws_co.append([])
        r_idx = ws_co.max_row
        is_zebra = idx_p % 2 == 0
        for col_idx, val in enumerate(row_data, start=2):
            cell = ws_co.cell(row=r_idx, column=col_idx, value=val)
            cell.font = body_font
            cell.border = thin_border
            if is_zebra:
                cell.fill = zebra_fill
            
            if col_idx in [3, 4]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left

# Ajusta largura das colunas da aba de coleções
for col in ws_co.columns:
    max_len = max(len(str(cell.value or '')) for cell in col)
    col_letter = get_column_letter(col[0].column)
    ws_co.column_dimensions[col_letter].width = max(max_len + 3, 10)


# Salva a planilha
output_path = os.path.abspath(os.path.join(os.getcwd(), "foc_serie_b_rodada_6.xlsx"))
wb.save(output_path)

print(f"Planilha Excel criada com sucesso em: {output_path}")
